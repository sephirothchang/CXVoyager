# SPDX-License-Identifier: GPL-3.0-or-later
# This file is part of CXVoyager.
#
# CXVoyager is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# CXVoyager is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with CXVoyager.  If not, see <https://www.gnu.org/licenses/>.

"""Excel 规划表解析与数据清洗。"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

from cxvoyager.common.system_constants import PLAN_KEYWORDS, PLAN_SHEETS
from cxvoyager.models import HostRow, MgmtInfo, PlanModel, VirtualNetworkRow
from . import field_variables as plan_vars

logger = logging.getLogger(__name__)


def find_plan_file(base_dir: Path) -> Path | None:
    """模糊匹配定位规划表文件。"""
    candidates = []
    for p in base_dir.glob("*.xlsx"):
        name = p.name
        if all(k in name for k in PLAN_KEYWORDS) and not name.startswith("~$"):
            candidates.append(p)
    if not candidates:
        return None
    # 选择修改时间最新的
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.replace("\n", " ").replace("\r", " ").strip()
        value = re.sub(r"\s+", " ", value)
    return value


def _resolve_value(raw: Any, default: Any) -> Any:
    if raw is None:
        return default
    cleaned = _clean_cell(raw)
    if cleaned in (None, ""):
        return default
    return cleaned


def _extract_variables(workbook: openpyxl.Workbook) -> Tuple[Dict[str, Any], List[str]]:
    values: Dict[str, Any] = {}
    missing_sheets: List[str] = []
    for key, var in plan_vars.PLAN_VARIABLES.items():
        if var.sheet not in workbook.sheetnames:
            if var.sheet not in missing_sheets:
                missing_sheets.append(var.sheet)
            values[key] = var.default
            continue
        ws = workbook[var.sheet]
        values[key] = _resolve_value(ws[var.cell].value, var.default)
    return values, missing_sheets


def _value(values: Dict[str, Any], ref: plan_vars.PlanVariable | str | None) -> Any:
    if ref is None:
        return None
    key = ref.key if isinstance(ref, plan_vars.PlanVariable) else ref
    return values.get(key)


def _build_virtual_network_records(values: Dict[str, Any]) -> List[Dict[str, Any]]:
    cluster_name = _value(values, plan_vars.CLUSTER_NAME) or _value(values, plan_vars.HOST_CLUSTER_NAME)
    records: List[Dict[str, Any]] = []
    mgmt_ports_value = _value(values, plan_vars.MGMT_SWITCH_PORTS)
    mgmt_bond_value = _value(values, plan_vars.MGMT_BOND_MODE)

    def add_record(
        label: str,
        vds_ref: plan_vars.PlanVariable | None,
        name_ref: plan_vars.PlanVariable | None,
        subnet_ref: plan_vars.PlanVariable | None,
        ports_ref: plan_vars.PlanVariable | str | List[str] | None,
        bond_ref: plan_vars.PlanVariable | str | None,
        vlan_ref: plan_vars.PlanVariable | None,
        vlan_type_ref: plan_vars.PlanVariable | None,
        gateway_ref: plan_vars.PlanVariable | None,
    ) -> None:
        def resolve(ref: plan_vars.PlanVariable | str | List[str] | None) -> Any:
            if isinstance(ref, plan_vars.PlanVariable):
                return _value(values, ref)
            if isinstance(ref, str) and ref in plan_vars.PLAN_VARIABLES:
                return _value(values, ref)
            return ref

        def _clean(value: Any) -> Any:
            if isinstance(value, str):
                stripped = value.strip()
                return stripped or None
            return value

        network_name = _clean(resolve(name_ref))
        subnet = _clean(resolve(subnet_ref))
        if not subnet:
            return
        record: Dict[str, Any] = {
            "网络标识": label,
            "集群名称": cluster_name,
            "虚拟交换机": resolve(vds_ref),
            "虚拟机网络": network_name,
            "subnetwork": subnet,
            "主机端口": resolve(ports_ref),
            "网口绑定模式": resolve(bond_ref),
            "vlan_id": resolve(vlan_ref),
            "vlan_type": resolve(vlan_type_ref),
            "gateway": resolve(gateway_ref),
        }
        records.append(record)

    add_record(
        "default",
        plan_vars.MGMT_VDS_NAME,
        plan_vars.DEFAULT_MGMT_NETWORK_NAME,
        plan_vars.DEFAULT_MGMT_SUBNET,
        mgmt_ports_value,
        mgmt_bond_value,
        plan_vars.DEFAULT_MGMT_VLAN_ID,
        plan_vars.DEFAULT_MGMT_VLAN_TYPE,
        plan_vars.DEFAULT_MGMT_GATEWAY,
    )
    add_record(
        "extra_mgmt",
        plan_vars.MGMT_VDS_NAME,
        plan_vars.EXTRA_NETWORK_NAME,
        plan_vars.EXTRA_SUBNET,
        mgmt_ports_value,
        mgmt_bond_value,
        plan_vars.EXTRA_VLAN_ID,
        plan_vars.EXTRA_VLAN_TYPE,
        plan_vars.EXTRA_GATEWAY,
    )
    add_record(
        "storage",
        plan_vars.STORAGE_VDS_NAME,
        plan_vars.STORAGE_NETWORK_NAME,
        plan_vars.STORAGE_SUBNET,
        plan_vars.STORAGE_SWITCH_PORTS,
        plan_vars.STORAGE_BOND_MODE,
        plan_vars.STORAGE_VLAN_ID,
        plan_vars.STORAGE_VLAN_TYPE,
        plan_vars.STORAGE_GATEWAY,
    )
    add_record(
        "backup",
        plan_vars.BACKUP_VDS_NAME,
        plan_vars.BACKUP_NETWORK_NAME,
        plan_vars.BACKUP_SUBNET,
        plan_vars.BACKUP_SWITCH_PORTS,
        plan_vars.BACKUP_BOND_MODE,
        plan_vars.BACKUP_VLAN_ID,
        plan_vars.BACKUP_VLAN_TYPE,
        plan_vars.BACKUP_GATEWAY,
    )

    business_vds = _value(values, plan_vars.BUSINESS_VDS_NAME)
    business_ports = _value(values, plan_vars.BUSINESS_SWITCH_PORTS)
    business_bond = _value(values, plan_vars.BUSINESS_BOND_MODE)
    for idx in range(plan_vars.BUSINESS_NETWORK_MAX):
        name = _value(values, plan_vars.BUSINESS_NETWORK_NAME_KEYS[idx])
        subnet = _value(values, plan_vars.BUSINESS_NETWORK_SUBNET_KEYS[idx])
        if isinstance(name, str):
            name = name.strip() or None
        if isinstance(subnet, str):
            subnet = subnet.strip() or None
        if not subnet:
            continue
        record = {
            "网络标识": f"business_{idx + 1:02d}",
            "集群名称": cluster_name,
            "虚拟交换机": business_vds,
            "虚拟机网络": name,
            "subnetwork": subnet,
            "主机端口": business_ports,
            "网口绑定模式": business_bond,
            "vlan_id": _value(values, plan_vars.BUSINESS_NETWORK_VLAN_ID_KEYS[idx]),
            "vlan_type": _value(values, plan_vars.BUSINESS_NETWORK_VLAN_TYPE_KEYS[idx]),
            "gateway": _value(values, plan_vars.BUSINESS_NETWORK_GATEWAY_KEYS[idx]),
        }
        records.append(record)

    return records


def _build_host_records(values: Dict[str, Any]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    cluster_name = _value(values, plan_vars.HOST_CLUSTER_NAME) or _value(values, plan_vars.CLUSTER_NAME)
    cluster_vip = _value(values, plan_vars.CLUSTER_VIP)

    for idx in range(plan_vars.HOST_MAX):
        mgmt_ip = _value(values, plan_vars.HOST_MGMT_IP_KEYS[idx])
        hostname = _value(values, plan_vars.HOST_HOSTNAME_KEYS[idx])
        bmc_ip = _value(values, plan_vars.HOST_BMC_IP_KEYS[idx])
        storage_ip = _value(values, plan_vars.HOST_STORAGE_IP_KEYS[idx])
        if not any([mgmt_ip, hostname, bmc_ip, storage_ip]):
            continue
        record = {
            "序号": idx + 1,
            "集群名称": cluster_name,
            "集群VIP": cluster_vip,
            "SMTX主机名": hostname,
            "管理地址": mgmt_ip,
            "存储地址": storage_ip,
            "带外地址": bmc_ip,
            "带外用户名": _value(values, plan_vars.HOST_BMC_USER_KEYS[idx]),
            "带外密码": _value(values, plan_vars.HOST_BMC_PASSWORD_KEYS[idx]),
            "主机SSH用户名": _value(values, plan_vars.HOST_SSH_USER_KEYS[idx]),
            "主机SSH密码": _value(values, plan_vars.HOST_SSH_PASSWORD_KEYS[idx]),
        }
        records.append(record)

    return records


def _build_mgmt_records(values: Dict[str, Any]) -> List[Dict[str, Any]]:
    cloudtower_ip = _value(values, plan_vars.CLOUDTOWER_IP)
    obs_ip = _value(values, plan_vars.OBS_IP)
    backup_ip = _value(values, plan_vars.BACKUP_IP)
    controller_ips = [
        _value(values, key) for key in plan_vars.ER_CONTROLLER_NODE_IP_KEYS if _value(values, key)
    ]
    other_components = [ip for ip in [obs_ip, backup_ip, *controller_ips] if ip]

    record = {
        "Cloudtower IP": cloudtower_ip,
        "root密码": _value(values, plan_vars.CLOUDTOWER_ROOT_PASSWORD),
        "其他组件IP": ",".join(other_components) if other_components else None,
        "OBS IP": obs_ip,
        "OBS 序列号": _value(values, plan_vars.OBS_SERIAL),
        "备份 IP": backup_ip,
        "备份 序列号": _value(values, plan_vars.BACKUP_SERIAL),
        "Cloudtower 序列号": _value(values, plan_vars.CLOUDTOWER_SERIAL),
        "ER 控制器集群VIP": _value(values, plan_vars.ER_CONTROLLER_CLUSTER_VIP),
        "ER 控制器节点IP列表": controller_ips,
        "ER 控制器序列号": _value(values, plan_vars.ER_CONTROLLER_SERIAL),
        "NTP 服务器": [
            _value(values, key) for key in plan_vars.NTP_SERVER_KEYS if _value(values, key)
        ],
        "DNS 服务器": [
            _value(values, key) for key in plan_vars.DNS_SERVER_KEYS if _value(values, key)
        ],
        "临时测试IP范围": {
            "start": _value(values, plan_vars.TEST_IP_RANGE_START),
            "end": _value(values, plan_vars.TEST_IP_RANGE_END),
        },
        "Cloudtower 组织名称": _value(values, plan_vars.CLOUDTOWER_ORGANIZATION_NAME),
        "Cloudtower 数据中心名称": _value(values, plan_vars.CLOUDTOWER_DATACENTER_NAME),
    }

    return [record]


def _build_hosts_extra(values: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "cluster_function": _value(values, plan_vars.CLUSTER_FUNCTION),
        "fisheye_admin_user": _value(values, plan_vars.FISHEYE_ADMIN_USER),
        "fisheye_admin_password": _value(values, plan_vars.FISHEYE_ADMIN_PASSWORD),
        "cluster_serial": _value(values, plan_vars.CLUSTER_SERIAL),
    }


def _build_derived_network(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    vds_map: Dict[str, Dict[str, Any]] = {}
    networks: List[Dict[str, Any]] = []
    for r in records:
        subnetwork = r.get("subnetwork")
        if not subnetwork:
            continue
        vds_name = r.get("虚拟交换机") or r.get("vds") or "vds0"
        nic_raw = r.get("主机端口") or ""
        if isinstance(nic_raw, str):
            nic_list = [n.strip() for n in re.split(r"[,/;\s]+", nic_raw) if n and n.strip()]
        elif isinstance(nic_raw, (list, tuple)):
            nic_list = [str(n).strip() for n in nic_raw if str(n).strip()]
        else:
            nic_list = []
        bond_mode_raw = r.get("网口绑定模式") or "active-backup"
        bond_mode = bond_mode_raw.strip().lower() if isinstance(bond_mode_raw, str) else "active-backup"
        bond_internal = {
            "active-backup": "ACTIVE_BACKUP",
            "balance-tcp": "LACP",
            "balance-slb": "SLB",
        }.get(bond_mode, "ACTIVE_BACKUP")
        if vds_name not in vds_map:
            vds_map[vds_name] = {
                "name": vds_name,
                "bond": {"mode": bond_internal, "nics": nic_list[:]},
                "mtu": 1500,
            }
        else:
            existing = vds_map[vds_name]["bond"].setdefault("nics", [])
            for nic in nic_list:
                if nic not in existing:
                    existing.append(nic)
            vds_map[vds_name]["bond"]["mode"] = bond_internal
        net_name = r.get("虚拟机网络") or r.get("network") or f"{vds_name}-net"
        networks.append(
            {
                "name": net_name,
                "vds": vds_name,
                "subnetwork": subnetwork,
                "metadata": {
                    "type": r.get("网络标识"),
                    "vlan_id": r.get("vlan_id"),
                    "vlan_type": r.get("vlan_type"),
                    "gateway": r.get("gateway"),
                },
            }
        )
    return {"vdses": list(vds_map.values()), "networks": networks}


def parse_plan(file_path: Path) -> Dict[str, Any]:
    """解析核心 sheet，返回结构化字典。"""
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    missing_declared = [sheet for sheet in PLAN_SHEETS.values() if sheet not in workbook.sheetnames]
    for sheet_name in missing_declared:
        logger.warning("缺少 sheet: %s", sheet_name)

    variables, missing_from_vars = _extract_variables(workbook)
    for sheet_name in missing_from_vars:
        if sheet_name not in missing_declared:
            logger.warning("变量读取时缺少 sheet: %s", sheet_name)

    virtual_network_records = _build_virtual_network_records(variables)
    host_records = _build_host_records(variables)
    mgmt_records = _build_mgmt_records(variables)

    data: Dict[str, Any] = {
        "variables": variables,
        "virtual_network": {"records": virtual_network_records},
        "hosts": {
            "records": host_records,
            "extra": _build_hosts_extra(variables),
        },
        "mgmt": {"records": mgmt_records},
        "_meta": {
            "source_file": str(file_path),
            "missing_sheets": sorted(set(missing_declared) | set(missing_from_vars)),
        },
    }

    try:
        data["_derived_network"] = _build_derived_network(virtual_network_records)
    except Exception as exc:  # noqa: BLE001
        logger.warning("派生网络结构失败: %s", exc)

    return data


def to_model(parsed: Dict[str, Any]) -> PlanModel:
    """将 parse_plan 的结果转换为强类型模型 PlanModel。"""
    def _build_vn() -> List[VirtualNetworkRow]:
        sec = parsed.get("virtual_network", {})
        recs = sec.get("records", []) if isinstance(sec, dict) else []
        out: List[VirtualNetworkRow] = []
        for r in recs:
            try:
                out.append(VirtualNetworkRow(**r))
            except Exception:  # noqa: broad
                continue
        return out

    def _build_hosts() -> List[HostRow]:
        sec = parsed.get("hosts", {})
        recs = sec.get("records", []) if isinstance(sec, dict) else []
        out: List[HostRow] = []
        for r in recs:
            try:
                out.append(HostRow(**r))
            except Exception:
                continue
        return out

    def _build_mgmt() -> MgmtInfo | None:
        sec = parsed.get("mgmt", {})
        recs = sec.get("records", []) if isinstance(sec, dict) else []
        if not recs:
            return None
        try:
            return MgmtInfo(**recs[0])
        except Exception as exc:  # noqa: BLE001
            logger.warning("规划表 mgmt 区域转换失败，使用空管理信息继续", exc_info=exc)
            return MgmtInfo()

    return PlanModel(
        virtual_network=_build_vn(),
        hosts=_build_hosts(),
        mgmt=_build_mgmt(),
        source_file=parsed.get("_meta", {}).get("source_file"),
    )

